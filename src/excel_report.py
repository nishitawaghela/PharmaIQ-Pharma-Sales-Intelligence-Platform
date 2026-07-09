import pandas as pd
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
import os
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

load_dotenv()
engine = create_engine(os.getenv('DATABASE_URL'))

# ── colour palette ──────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
GREEN       = "70AD47"
RED         = "FF0000"
YELLOW      = "FFD966"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F2F2F2"

def thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def fetch_data():
    sales = pd.read_sql(
        "SELECT * FROM sales_performance", engine)
    reps  = pd.read_sql("SELECT * FROM reps", engine)
    mkt   = pd.read_sql("SELECT * FROM market_share", engine)
    return sales, reps, mkt

# ── Sheet 1: Executive KPI Summary ──────────────────────────────
def create_kpi_sheet(ws, sales, reps):
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = DARK_BLUE

    # header banner
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "PharmaIQ — Executive Sales Intelligence Report"
    cell.font      = Font(name="Calibri", size=16,
                          bold=True, color=WHITE)
    cell.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center",
                               vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    cell = ws["A2"]
    cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    cell.font      = Font(name="Calibri", size=10,
                          italic=True, color=WHITE)
    cell.fill      = PatternFill("solid", fgColor=MED_BLUE)
    cell.alignment = Alignment(horizontal="center")

    # KPI calculations
    total_sales    = sales["actual_sales"].sum()
    total_quota    = sales["quota"].sum()
    overall_att    = (total_sales / total_quota) * 100
    avg_att        = sales["attainment_pct"].mean()
    top_region     = (sales.groupby("region")["actual_sales"]
                      .sum().idxmax())
    top_drug       = (sales.groupby("drug")["actual_sales"]
                      .sum().idxmax())
    reps_above     = (sales.groupby("rep_id")["attainment_pct"]
                      .mean().gt(100).sum())
    reps_below     = (sales.groupby("rep_id")["attainment_pct"]
                      .mean().lt(80).sum())

    kpis = [
        ("Total Revenue",        f"${total_sales:,.0f}",  GREEN),
        ("Total Quota",          f"${total_quota:,.0f}",  MED_BLUE),
        ("Overall Attainment",   f"{overall_att:.1f}%",   GREEN if overall_att >= 100 else YELLOW),
        ("Avg Rep Attainment",   f"{avg_att:.1f}%",       GREEN if avg_att >= 100 else YELLOW),
        ("Top Region",           top_region,               MED_BLUE),
        ("Top Drug",             top_drug,                 MED_BLUE),
        ("Reps Above Quota",     str(reps_above),          GREEN),
        ("Reps Below 80% Quota", str(reps_below),          RED),
    ]

    ws["A4"].value = "KEY PERFORMANCE INDICATORS"
    ws["A4"].font  = Font(bold=True, size=12, color=WHITE)
    ws["A4"].fill  = PatternFill("solid", fgColor=MED_BLUE)
    ws.merge_cells("A4:F4")
    ws["A4"].alignment = Alignment(horizontal="center")

    for i, (label, value, colour) in enumerate(kpis):
        row = 5 + i
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font  = Font(bold=True)
        ws.cell(row=row, column=1).fill  = PatternFill(
            "solid", fgColor=LIGHT_GREY)
        ws.cell(row=row, column=1).border = thin_border()

        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=2).font  = Font(
            bold=True, color=colour)
        ws.cell(row=row, column=2).border = thin_border()
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

# ── Sheet 2: Rep Performance Table ──────────────────────────────
def create_rep_sheet(ws, sales, reps):
    ws.title = "Rep Performance"
    ws.sheet_properties.tabColor = MED_BLUE

    rep_summary = (sales.groupby("rep_id")
                   .agg(total_quota   =("quota",          "sum"),
                        total_sales   =("actual_sales",   "sum"),
                        avg_attainment=("attainment_pct", "mean"),
                        total_visits  =("total_visits",   "sum"))
                   .reset_index())
    rep_summary = rep_summary.merge(
        reps[["rep_id","name","region","territory","drug_promoted"]],
        on="rep_id")
    rep_summary = rep_summary.sort_values(
        "avg_attainment", ascending=False)

    headers = ["Rep ID","Name","Region","Territory",
               "Drug","Total Quota","Total Sales",
               "Avg Attainment %","Total Visits","Status"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value     = h
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin_border()

    for row_idx, row in enumerate(
            rep_summary.itertuples(), 2):
        att    = row.avg_attainment
        status = ("🟢 On Track"  if att >= 100 else
                  "🟡 At Risk"   if att >= 80  else
                  "🔴 Critical")
        colour = (GREEN  if att >= 100 else
                  YELLOW if att >= 80  else RED)

        vals = [row.rep_id, row.name, row.region,
                row.territory, row.drug_promoted,
                f"${row.total_quota:,.0f}",
                f"${row.total_sales:,.0f}",
                f"{att:.1f}%",
                row.total_visits, status]

        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value  = val
            cell.border = thin_border()
            cell.fill   = PatternFill(
                "solid",
                fgColor=LIGHT_GREY if row_idx % 2 == 0
                        else WHITE)
            if col_idx == 9:          # attainment col
                cell.font = Font(bold=True, color=colour)

    col_widths = [10,22,12,14,14,14,14,18,13,14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 3: Regional Sales Bar Chart ───────────────────────────
def create_regional_chart(ws, sales):
    ws.title = "Regional Analysis"
    ws.sheet_properties.tabColor = GREEN

    region_data = (sales.groupby("region")["actual_sales"]
                   .sum().reset_index())
    region_data.columns = ["Region","Total Sales"]

    ws["A1"].value = "Region"
    ws["B1"].value = "Total Sales"
    ws["A1"].font  = Font(bold=True, color=WHITE)
    ws["B1"].font  = Font(bold=True, color=WHITE)
    ws["A1"].fill  = PatternFill("solid", fgColor=DARK_BLUE)
    ws["B1"].fill  = PatternFill("solid", fgColor=DARK_BLUE)

    for i, row in enumerate(region_data.itertuples(), 2):
        ws.cell(row=i, column=1).value = row.Region
        ws.cell(row=i, column=2).value = round(row._2, 2)

    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Total Sales by Region"
    chart.y_axis.title = "Sales ($)"
    chart.x_axis.title = "Region"
    chart.style   = 10
    chart.width   = 20
    chart.height  = 15

    data = Reference(ws, min_col=2, min_row=1,
                     max_row=len(region_data)+1)
    cats = Reference(ws, min_col=1, min_row=2,
                     max_row=len(region_data)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")

# ── Sheet 4: Drug Market Share ───────────────────────────────────
def create_drug_sheet(ws, sales):
    ws.title = "Drug Performance"
    ws.sheet_properties.tabColor = YELLOW

    drug_data = (sales.groupby("drug")
                 .agg(total_sales   =("actual_sales",   "sum"),
                      avg_attainment=("attainment_pct", "mean"))
                 .reset_index()
                 .sort_values("total_sales", ascending=False))

    headers = ["Drug","Total Sales","Avg Attainment %","Rank"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.font  = Font(bold=True, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=DARK_BLUE)
        cell.border = thin_border()

    for i, row in enumerate(drug_data.itertuples(), 2):
        ws.cell(row=i, column=1).value = row.drug
        ws.cell(row=i, column=2).value = f"${row.total_sales:,.0f}"
        ws.cell(row=i, column=3).value = f"{row.avg_attainment:.1f}%"
        ws.cell(row=i, column=4).value = i - 1
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = thin_border()
            ws.cell(row=i, column=col).fill   = PatternFill(
                "solid",
                fgColor=LIGHT_GREY if i % 2 == 0 else WHITE)

    for i, w in enumerate([18,18,18,8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 5: Monthly Trend Line Chart ───────────────────────────
def create_trend_sheet(ws, sales):
    ws.title = "Monthly Trends"
    ws.sheet_properties.tabColor = MED_BLUE

    monthly = (sales.groupby("month")["actual_sales"]
               .sum().reset_index()
               .sort_values("month"))

    ws["A1"].value = "Month"
    ws["B1"].value = "Total Sales"
    for cell in [ws["A1"], ws["B1"]]:
        cell.font  = Font(bold=True, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=DARK_BLUE)

    for i, row in enumerate(monthly.itertuples(), 2):
        ws.cell(row=i, column=1).value = row.month
        ws.cell(row=i, column=2).value = round(row.actual_sales, 2)

    chart = LineChart()
    chart.title        = "Monthly Revenue Trend"
    chart.y_axis.title = "Sales ($)"
    chart.x_axis.title = "Month"
    chart.style        = 10
    chart.width        = 25
    chart.height       = 15

    data = Reference(ws, min_col=2, min_row=1,
                     max_row=len(monthly)+1)
    cats = Reference(ws, min_col=1, min_row=2,
                     max_row=len(monthly)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")

# ── Main ─────────────────────────────────────────────────────────
def generate_report():
    print("Fetching data from Supabase...")
    sales, reps, mkt = fetch_data()

    wb = openpyxl.Workbook()

    # Sheet 1
    create_kpi_sheet(wb.active, sales, reps)
    # Sheet 2
    create_rep_sheet(wb.create_sheet(), sales, reps)
    # Sheet 3
    create_regional_chart(wb.create_sheet(), sales)
    # Sheet 4
    create_drug_sheet(wb.create_sheet(), sales)
    # Sheet 5
    create_trend_sheet(wb.create_sheet(), sales)

    os.makedirs("reports", exist_ok=True)
    filename = (f"reports/PharmaIQ_Report_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(filename)
    print(f"Report saved: {filename}")
    return filename

if __name__ == '__main__':
    generate_report()